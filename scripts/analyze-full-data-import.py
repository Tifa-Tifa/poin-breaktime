from collections import Counter
from datetime import datetime
import json
import os
from openpyxl import load_workbook

root = os.path.dirname(os.path.dirname(__file__))
with open(os.path.join(root, "data", "db.json"), encoding="utf-8") as f:
    db = json.load(f)

employees = {str(e["name"]).strip().upper(): e for e in db["employees"]}
inactive = {name for name, e in employees.items() if e.get("status") == "INACTIVE"}
captains = {name for name, e in employees.items() if e.get("position") == "KAPTEN"}

path = os.path.join(os.environ["TEMP"], "poin-breaktime-full.xlsx")
ws = load_workbook(path, data_only=True, read_only=True)["DATA"]

rows = []
for values in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    date, name, description, qty, outlet, position, gender, category, point, *_ = values
    if not date or not name or not description or point is None:
        continue
    key = str(name).strip().upper()
    rows.append((date, key, str(description).strip(), qty, point))

counts = Counter(name for _, name, *_ in rows)
matched_active = [r for r in rows if r[1] in employees and r[1] not in inactive and r[1] not in captains]
skipped_inactive = [r for r in rows if r[1] in inactive]
skipped_captains = [r for r in rows if r[1] in captains]
unmatched = [r for r in rows if r[1] not in employees]
dates = [r[0] for r in rows if isinstance(r[0], datetime)]

print("VALID_ROWS", len(rows))
print("DATE_RANGE", min(dates).date(), max(dates).date())
print("DISTINCT_NAMES", len(counts))
print("MATCHED_ACTIVE_ROWS", len(matched_active), "NAMES", len({r[1] for r in matched_active}))
print("SKIPPED_INACTIVE_ROWS", len(skipped_inactive), "NAMES", sorted({r[1] for r in skipped_inactive}))
print("SKIPPED_CAPTAIN_ROWS", len(skipped_captains), "NAMES", sorted({r[1] for r in skipped_captains}))
print("UNMATCHED_ROWS", len(unmatched), "NAMES", len({r[1] for r in unmatched}))
print("UNMATCHED_DETAIL")
for name, count in Counter(r[1] for r in unmatched).most_common():
    print(name, count)

print("INACTIVE_DETAIL")
for name, count in Counter(r[1] for r in skipped_inactive).most_common():
    print(name, count)

print("YEAR_MONTH_COUNTS")
for month, count in sorted(Counter(r[0].strftime("%Y-%m") for r in rows if isinstance(r[0], datetime)).items()):
    print(month, count)

rules = {str(r["description"]).strip().casefold() for r in db["rules"]}
unmatched_rules = Counter(r[2] for r in matched_active if r[2].casefold() not in rules)
print("UNMATCHED_RULE_ROWS", sum(unmatched_rules.values()), "DESCRIPTIONS", len(unmatched_rules))
for description, count in unmatched_rules.most_common():
    print(description, count)
