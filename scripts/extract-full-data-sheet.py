import json
import os
from datetime import date, datetime

from openpyxl import load_workbook

source = os.path.join(os.environ["TEMP"], "poin-breaktime-full.xlsx")
target = os.path.join(os.environ["TEMP"], "poin-breaktime-full-data.json")
ws = load_workbook(source, data_only=True, read_only=True)["DATA"]

records = []
for values in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    tanggal, nama, deskripsi, kelipatan, outlet, jabatan, gender, kategori, poin, *_ = values
    if not tanggal or not nama or not deskripsi or poin is None:
        continue
    if isinstance(tanggal, (datetime, date)):
        tanggal = tanggal.strftime("%Y-%m-%d")
    records.append({
        "date": str(tanggal),
        "name": str(nama).strip(),
        "description": str(deskripsi).strip(),
        "quantity": kelipatan,
        "sourceOutlet": outlet,
        "sourcePosition": jabatan,
        "sourceGender": gender,
        "category": kategori,
        "points": float(poin),
    })

with open(target, "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, separators=(",", ":"))

print(json.dumps({"target": target, "rows": len(records)}))
