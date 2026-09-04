from collections import Counter, defaultdict
import os
from openpyxl import load_workbook

path = os.path.join(os.environ["TEMP"], "poin-breaktime-full.xlsx")
wb = load_workbook(path, data_only=True, read_only=False)
print("SHEETS", wb.sheetnames)
ws = wb["DATA"]
print("DATA_DIM", ws.max_row, ws.max_column, "FILTER", ws.auto_filter.ref)
print("HIDDEN_ROWS", sum(1 for i in range(1, ws.max_row + 1) if ws.row_dimensions[i].hidden))

for r in range(1, min(ws.max_row, 20) + 1):
    print("ROW", r, [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)])

# Inventory non-empty columns and font colors to locate employee-name formatting.
counts = Counter()
samples = defaultdict(list)
for row in ws.iter_rows():
    for cell in row:
        if cell.value not in (None, ""):
            color = cell.font.color
            if color is None:
                key = "none"
            else:
                key = f"{color.type}:{color.rgb or color.indexed or color.theme}"
            counts[(cell.column, key)] += 1
            if len(samples[(cell.column, key)]) < 5:
                samples[(cell.column, key)].append(str(cell.value))

for (col, color), count in sorted(counts.items()):
    if color != "none":
        print("COLOR", col, color, count, samples[(col, color)])

print("SHEET_COLOR_INVENTORY")
for sheet in wb.worksheets:
    inventory = Counter()
    inventory_samples = defaultdict(list)
    styled_names = defaultdict(set)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            color = cell.font.color
            rgb = color.rgb if color is not None and color.type == "rgb" else None
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else None
            key = (cell.style_id, rgb, fill)
            inventory[key] += 1
            if len(inventory_samples[key]) < 5:
                inventory_samples[key].append(str(cell.value))
    print("SHEET", sheet.title, sheet.max_row, sheet.max_column, "CF", len(sheet.conditional_formatting))
    for key, count in inventory.most_common():
        style_id, rgb, fill = key
        if rgb not in (None, "FF000000", "FF434343") or fill not in (None, "00000000", "FFFFFFFF"):
            print("STYLE", sheet.title, key, count, inventory_samples[key])
    for cf in sheet.conditional_formatting:
        print("CF_RANGE", sheet.title, str(cf))
        for rule in sheet.conditional_formatting[cf]:
            print("CF_RULE", rule.type, getattr(rule, "formula", None), getattr(rule, "operator", None), getattr(rule, "dxfId", None))
